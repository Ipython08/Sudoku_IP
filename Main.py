import copy
import random
import openpyxl
import json
import os
from openpyxl.styles import Border, Side, Alignment, Font
import mysql.connector
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

mysql_connector = mysql.connector

DB_CONFIG = {
    'host': 'localhost',
    'user': 'sudoku',
    'password': 'abc',
    'database': 'sudoku_games_db'
}

opboard = [[0 for _ in range(9)] for _ in range(9)]

def find_empty(opboard):
    """Finds an empty cell (represented by 0) on the board."""
    for i in range(9):
        for j in range(9):
            if opboard[i][j] == 0:
                return (i, j)  # row, col
    return None


def is_valid(num, pos):
    """
    Checks if placing a number in a given position is valid.
        num (int): The number to place.
        pos (tuple): The (row, col) position to check.
    Returns bool True if valid, False otherwise.
    """
    row, col = pos

    # Check row
    for j in range(9):
        if opboard[row][j] == num and col != j:
            return False

    # Check column
    for i in range(9):
        if opboard[i][col] == num and row != i:
            return False

    # Check 3x3 box
    box_start_row, box_start_col = 3 * (row // 3), 3 * (col // 3)
    for i in range(box_start_row, box_start_row + 3):
        for j in range(box_start_col, box_start_col + 3):
            if opboard[i][j] == num and (i, j) != pos:
                return False

    return True


def solve():
    """
    Solves the Sudoku puzzle using a recrusive algorithm.
    Modifies the board in place.
    Returns:
        bool: True if a solution is found, False otherwise.
    """
    global opboard
    find = find_empty(opboard)
    if not find:
        return True  # Puzzle is solved
    else:
        row, col = find

    for num in range(1, 10):
        opboard[row][col] = num
        if is_valid(num, (row, col)):
            if solve():
                return True

        opboard[row][col] = 0  # Backtrack (reset to empty)

    return False


def generate(difficulty):

    #Generates a new Sudoku puzzle.

    # Reset the board
    global opboard
    opboard = [[0 for _ in range(9)] for _ in range(9)]

    # 1. Fill the diagonal boxes
    for i in range(0, 9, 3):
        nums = list(range(1, 10))
        random.shuffle(nums)
        for r in range(3):
            for c in range(3):
                opboard[i + r][i + c] = nums.pop()

    # 2. Solve the full board to guarantee a solution (this fills the rest)

    solve()

    # 3. Remove numbers to create the puzzle

    removals = min(difficulty, 64) #Max removals to make a solveable puzzle is 64

    count = 0
    cells = [(r, c) for r in range(9) for c in range(9)]
    random.shuffle(cells)

    for row, col in cells:
        if count >= removals:
            break
        if opboard[row][col] != 0:
            opboard[row][col] = 0
            count += 1

def display():
    board_str = "\n"
    for i in range(9):
        if i % 3 == 0 and i != 0:
            board_str += "---------------------\n"
        for j in range(9):
            if j % 3 == 0 and j != 0:
                board_str += "| "
            cell = opboard[i][j]
            # Use '.' for empty cells in the console output
            board_str += f"{cell if cell != 0 else '.'} "
        board_str += "\n"
    print(board_str)
    os.startfile("Puzzle.xlsx")

def export_to_excel(filename="Puzzle.xlsx"):

    #Exports the current puzzle board to a formatted Excel file.

    # Define styles
    thin_border = Side(style='thin',)
    thick_border = Side(style='medium',)
    default_font = Font(name='Arial', size=14, bold=True)
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sudoku Puzzle"
    # Set column width and row height
    col_width = 4
    row_height = 25
    for i in range(1,10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_width
        ws.row_dimensions[i].height = row_height
    # Loop through the board and write values
    for r in range(9):
        for c in range(9):
            cell_value = opboard[r][c]
            # Write number or empty string if 0 (blank cell)
            ws.cell(row=r + 1, column=c + 1, value=cell_value if cell_value != 0 else "")
            cell = ws.cell(row=r + 1, column=c + 1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = default_font
            # Define cell borders
            top = thin_border
            bottom = thin_border
            left = thin_border
            right = thin_border
            # Apply thick borders for the 3x3 boxes
            if r % 3 == 0:
                top = thick_border
            if (r + 1) % 3 == 0:
                bottom = thick_border
            if c % 3 == 0:
                left = thick_border
            if (c + 1) % 3 == 0:
                right = thick_border
            # Apply thick borders to the outer edge of the entire grid
            if r == 0: top = thick_border
            if r == 8: bottom = thick_border
            if c == 0: left = thick_border
            if c == 8: right = thick_border
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    wb.save(filename)
def _load_board_from_excel(filename):

    global opboard

    if not os.path.exists(filename):
        return None, f"File not found: {filename}"
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        loaded_board = []
        for r in range(1, 10):
            row_data = []
            for c in range(1, 10):
                cell_value = ws.cell(row=r, column=c).value
                # Convert to integer, treating None/empty strings as 0
                try:
                    num = int(cell_value) if cell_value not in [None, ""] else 0
                    # Validate range (1 to size, or 0 for empty)
                    if not (1 <= num <= 9 or num == 0):
                        return None, f"Invalid number {cell_value} found in cell ({r}, {c}). Expected 1-9 or empty."
                    row_data.append(num)
                except (ValueError, TypeError):
                    # Handles cases where cell_value is non-numeric text
                    return None, f"Cell ({r}, {c}) contains non-numeric data: '{cell_value}'"
            loaded_board.append(row_data)

        opboard=loaded_board
        export_to_excel("Puzzle.xlsx")
        return loaded_board

    except Exception as e:
        return None, f"Error loading or parsing Excel file: {e}"

def verify_solution_from_excel(filename):
    """
    Loads a board from an Excel file and verifies if it is a complete and
    valid Sudoku solution.
    Returns: True if the board is a complete and valid solution, False otherwise.
    """
    loaded_board= _load_board_from_excel(filename)
    if loaded_board is None:
        print(f"\nVerification failed...")
        return False
    # Expected set of numbers for a valid row, column, or box
    valid_set = set(range(1, 10))
    # 1. Check for completeness (no empty cells/0s)
    for r in range(9):
        if 0 in loaded_board[r]:
            print("Verification failed... Board is incomplete (contains empty cells).")
            return False
    # 2. Check all rows, columns, and boxes
    # Check rows
    for r in range(9):
        if set(loaded_board[r]) != valid_set:
            print(f"Verification failed... Row {r + 1} contains duplicates or invalid numbers.")
            return False
    # Check columns
    for c in range(9):
        col_list = [loaded_board[r][c] for r in range(9)]
        if set(col_list) != valid_set:
            print(f"Verification failed... Column {c + 1} contains duplicates or invalid numbers.")
            return False
    # Check boxes
    for box_r_start in range(0, 9, 3):
        for box_c_start in range(0, 9, 3):
            box_list = []
            for r in range(box_r_start, box_r_start + 3):
                for c in range(box_c_start, box_c_start + 3):
                    box_list.append(loaded_board[r][c])
            if set(box_list) != valid_set:
                print(
                    f"Verification failed... Box starting at ({box_r_start + 1}, {box_c_start + 1}) contains duplicates or invalid numbers.")
                return False
    print("Verification success... The board loaded from Excel is a complete and valid Sudoku solution.")
    return True

def save_game_to_db(user: str, game_name: str):

    #Saves the current board state to the MySQL database.

    if mysql_connector is None:
        print("\nMySQL connector is not available.")
        return
    # 1. Serialize the board (2D list of integers) into a JSON string
    board_json = json.dumps(opboard)
    # 2. Establish connection and ensure table exists
    conn = None
    try:
        conn = mysql_connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        # Create the games table if it does not exist
        create_table_query = """
                            CREATE TABLE IF NOT EXISTS sudoku_games (
                                -- Unique identifier for each game session.
                            id INT AUTO_INCREMENT PRIMARY KEY,
                        
                            -- Identifier for the user who owns the game record.
                            user_id VARCHAR(255) NOT NULL,
                        
                            -- A custom name the user might give the saved game.
                            game_name VARCHAR(255) NOT NULL,
                        
                            -- The complete state of the Sudoku board (e.g., a serialized string or JSON).
                            board_state TEXT NOT NULL,
                        
                            -- Timestamp for when the game was created or first saved.
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                            );
                             """
        cursor.execute(create_table_query)
        # 3. Insert the game data
        insert_query = """
                       INSERT INTO sudoku_games
                           (user_id, game_name, board_state)
                       VALUES (%s, %s, %s) \
                       """
        data = (user, game_name, board_json)
        cursor.execute(insert_query, data)
        conn.commit()
        print(f"\nSaved '{game_name}' for user {user} (ID: {cursor.lastrowid}).")
    except mysql_connector.Error as err:
        print(f"\nFailed to save game: {err}. Check DB_CONFIG and connection.")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

def list_games_from_db():
    #Fetches and prints all saved games for the current user from the database.
    global user

    if mysql_connector is None:
        print("\nMySQL connector is not available. Check installation and configuration.")
        return

    conn = None
    try:
        conn = mysql_connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        select_query = """
                       SELECT id, game_name, created_at
                       FROM sudoku_games
                       WHERE user_id= %s
                       ORDER BY created_at DESC \
                       """
        cursor.execute(select_query, (user,))

        results = cursor.fetchall()

        if not results:
            print(f"\nNo saved games found for user {user}.")
            return

        print(f"\n--- Saved Sudoku Games for User {user} ({len(results)} total) ---")

        # Define header and column widths for neat console printing
        header = ["ID", "Game Name", "Saved At"]
        widths = [5, 30, 25]

        # Print header
        header_line = f"{header[0]:<{widths[0]}} | {header[1]:<{widths[1]}} | {header[2]:<{widths[2]}}"
        print("-" * len(header_line))
        print(header_line)
        print("-" * len(header_line))

        # Print data rows
        for row in results:
            game_id, name, created_at = row
            # Format timestamp nicely
            formatted_date = created_at.strftime("%Y-%m-%d %H:%M:%S")

            data_line = (
                f"{game_id:<{widths[0]}} | "
                f"{name:<{widths[1]}} | "
                f"{formatted_date:<{widths[2]}}"
            )

            print(data_line)

        print("-" * len(header_line))


    except mysql_connector.Error as err:
        print(f"\nFailed to retrieve games: {err}. Please ensure the database is running and accessible.")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

def load_game_from_db(game_id):
    global user
    global opboard

    if mysql_connector is None:
        print("\nMySQL connector is not available. Check installation and configuration.")
        return

    try:
        if not game_id.isdigit():
            print("Invalid ID. Please enter a numerical game ID.")
            return
    except ValueError:
        print("Invalid input.")
        return

    conn = None
    try:
        conn = mysql_connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Retrieve the board state and size for the specific game ID and user
        select_query = """
                       SELECT board_state
                       FROM sudoku_games
                       WHERE id = %s 
                         AND user_id = %s 
                       """
        cursor.execute(select_query, (int(game_id), user))

        result = cursor.fetchone()

        if not result:
            print(f"\nGame with ID {game_id} not found for user {user}.")
            return

        board_json = result[0]

        # Deserialize the board state
        loaded_board = json.loads(board_json)

        # Create a new Sudoku instance and load the board
        puzzle = loaded_board
        opboard = copy.deepcopy(puzzle)
        export_to_excel("Puzzle.xlsx")
        print(f"\nSuccessfully loaded game ID {game_id}.")
        display()

    except mysql_connector.Error as err:
        print(f"\nFailed to load game: {err}. Please ensure the database is running and accessible.")
    except Exception as e:
        print(f"\nAn unexpected error occurred while loading: {e}")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

def delete_game_from_db():
    global user

    if mysql_connector is None:
        print("\nMySQL connector is not available. Check installation and configuration.")
        return

    try:
        game_id = input("Enter the ID of the game to delete (use Option 6 to view IDs): ").strip()
        if not game_id.isdigit():
            print("Invalid ID. Please enter a numerical game ID.")
            return
    except ValueError:
        print("Invalid input.")
        return

    conn = None
    try:
        conn = mysql_connector.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Delete the board for the specific game id and user
        delete_query = """
                       DELETE 
                       FROM sudoku_games
                       WHERE id = %s 
                         AND user_id = %s 
                       """
        cursor.execute(delete_query, (int(game_id), user))
        conn.commit()
        print(f"\nSuccessfully deleted game ID {game_id}.")

    except mysql_connector.Error as err:
        print(f"\nFailed to load game: {err}. Please ensure the database is running and accessible.")
    except Exception as e:
        print(f"\nAn unexpected error occurred while loading: {e}")
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

print("Welcome to Sudoku. Type 'help' to see available commands.")
user=input("Username: ")

while(True):
    
    command=input("Enter command: ").strip()
    
    if command == "help":
        print("\n" + "=" * 40)
        print("SUDOKU TOOL MENU")
        print("=" * 40)
        print("1. Generate New Puzzle")
        print("2. Print Current Board")
        print("3. Solve Current Puzzle")
        print("4. Verify Solution from Excel File")
        print("5. Save Current Board to Database")
        print("6. List All Saved Games from DB")
        print("7. Load Game from DB by ID")
        print("8. Load Game from an Excel file")
        print("9. Delete Game from DB by ID")
        print("10. Switch User ID")
        print("0. Exit")
        print("-" * 40)
    elif command == "1":
        d=input("Enter number of boxes to be empty (Max:64, Higher is harder) (Default:50)").strip()
        generate(int(d) if d!='' else 50)
        puzzle=copy.deepcopy(opboard)
        display()
        export_to_excel("Puzzle.xlsx")
        os.startfile("Puzzle.xlsx")
    elif command == "2":
        print("Current Board:")
        opboard = copy.deepcopy(puzzle)
        display()
    elif command == "3":
        print("Solved Board:")
        opboard = copy.deepcopy(puzzle)
        solve()
        display()
    elif command == "4":
        verify_solution_from_excel('Puzzle.xlsx')
        c=input("Save puzzle to db? (y/n): ").lower()
        if c == "y":
            nam=input("Game name: ").strip()
            opboard = copy.deepcopy(puzzle)
            save_game_to_db(user, nam)
    elif command == "5":
        nam = input("Game name: ").strip()
        opboard = copy.deepcopy(puzzle)
        save_game_to_db(user, nam)
    elif command == "6":
        list_games_from_db()
    elif command == "7":
        id = input("Enter the ID of the game to load (use Option 6 to view IDs): ").strip()
        load_game_from_db(id)
    elif command == "8":
        file_path = filedialog.askopenfilename(
            initialdir=os.getcwd(),  # Start the dialog in the current working directory
            title="Select a File",
            filetypes=(("Excel Sheets", "*.xlsx"), ("All files", "*.*"))
        )
        _load_board_from_excel(file_path)
        puzzle = copy.deepcopy(opboard)
        print('Loaded board from Excel sheet')
        display()
    elif command == "9":
        delete_game_from_db()
    elif command == "10":
        user = input("Username: ")
    elif command == "0":
        print("\nExiting Sudoku. Goodbye!")

        break
